from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import urllib.request
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.backfill_yahoo_generic_etf_names import merge_metadata_updates
from scripts.rebuild_dataset import TICKERS_CSV, normalize_sector


B3_SECTOR_ZIP_URL = "https://bvmf.bmfbovespa.com.br/InstDados/InformacoesEmpresas/ClassifSetorial.zip"
B3_INDUSTRY_XLSX_URL = (
    "https://sistemaswebb3-listados.b3.com.br/listedCompaniesProxy/CompanyCall/"
    "GetDownloadIndustryClassification/eyJsYW5ndWFnZSI6ImVuLXVzIiwicGFnZU51bWJlciI6MSwicGFnZVNpemUiOjEwMH0="
)
DEFAULT_OUTPUT_DIR = ROOT / "data" / "b3_verification"
DEFAULT_REPORT_JSON = DEFAULT_OUTPUT_DIR / "sector_classification_backfill.json"
DEFAULT_REPORT_CSV = DEFAULT_OUTPUT_DIR / "sector_classification_backfill.csv"
DEFAULT_METADATA_UPDATES_CSV = ROOT / "data" / "review_overrides" / "metadata_updates.csv"

B3_SECTOR_MAP = {
    "Petróleo, Gás e Biocombustíveis": "Energy",
    "Materiais Básicos": "Materials",
    "Bens Industriais": "Industrials",
    "Consumo não Cíclico": "Consumer Staples",
    "Consumo Cíclico": "Consumer Discretionary",
    "Saúde": "Health Care",
    "Tecnologia da Informação": "Information Technology",
    "Comunicações": "Communication Services",
    "Utilidade Pública": "Utilities",
    "Financeiro": "Financials",
}
B3_ENGLISH_SECTOR_MAP = {
    "Oil,Gas and Biofuels": "Energy",
    "Basic Materials": "Materials",
    "Capital Goods and Services": "Industrials",
    "Consumer Discretionary": "Consumer Discretionary",
    "Consumer Staples": "Consumer Staples",
    "Health Care": "Health Care",
    "Information Technology": "Information Technology",
    "Communications": "Communication Services",
    "Utilities": "Utilities",
    "Financial": "Financials",
}

REPORT_FIELDNAMES = [
    "ticker",
    "exchange",
    "asset_type",
    "name",
    "b3_code",
    "b3_name",
    "b3_sector_pt",
    "b3_subsector",
    "b3_segment",
    "sector_update",
    "decision",
]


@dataclass(frozen=True)
class B3SectorRow:
    code: str
    name: str
    sector_pt: str
    sector: str
    subsector: str
    segment: str


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def download_b3_sector_zip(url: str, timeout_seconds: float) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "free-ticker-database/3.0"})
    with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
        return response.read()


def download_b3_industry_xlsx(url: str, timeout_seconds: float) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "free-ticker-database/3.0"})
    with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
        return response.read()


def extract_workbook(zip_bytes: bytes) -> bytes:
    with ZipFile(BytesIO(zip_bytes)) as archive:
        xlsx_names = [name for name in archive.namelist() if name.lower().endswith(".xlsx")]
        if len(xlsx_names) != 1:
            raise ValueError(f"Expected one xlsx in B3 classification ZIP, found {xlsx_names!r}")
        return archive.read(xlsx_names[0])


def load_b3_sector_rows(workbook_bytes: bytes) -> list[B3SectorRow]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    worksheet = workbook.active
    current_sector = ""
    current_subsector = ""
    current_segment = ""
    rows: list[B3SectorRow] = []

    for values in worksheet.iter_rows(values_only=True):
        padded = list(values) + [None] * 7
        sector_cell, subsector_cell, segment_or_name_cell, code_cell = padded[:4]
        sector_text = str(sector_cell).strip() if sector_cell is not None else ""
        subsector_text = str(subsector_cell).strip() if subsector_cell is not None else ""
        segment_or_name = str(segment_or_name_cell).strip() if segment_or_name_cell is not None else ""
        code = str(code_cell).strip().upper() if code_cell is not None else ""

        if sector_text in B3_SECTOR_MAP:
            current_sector = sector_text
            current_subsector = subsector_text
            current_segment = segment_or_name
            continue
        if sector_text and sector_text != "SETOR ECONÔMICO":
            current_sector = ""
            current_subsector = ""
            current_segment = ""
            continue
        if subsector_text and not code and current_sector:
            current_subsector = subsector_text
            current_segment = segment_or_name
            continue
        if segment_or_name and not code and current_sector and segment_or_name.upper() != "SEGMENTO":
            current_segment = segment_or_name
            continue
        if not segment_or_name or not code or code in {"CÓDIGO", "CODIGO"}:
            continue

        sector = normalize_sector(B3_SECTOR_MAP.get(current_sector, ""), "Stock")
        if not sector:
            continue
        rows.append(
            B3SectorRow(
                code=code,
                name=segment_or_name,
                sector_pt=current_sector,
                sector=sector,
                subsector=current_subsector,
                segment=current_segment,
            )
        )
    return rows


def load_b3_industry_download_rows(workbook_bytes: bytes) -> list[B3SectorRow]:
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    worksheet = workbook.active
    current_sector = ""
    current_subsector = ""
    current_segment = ""
    rows: list[B3SectorRow] = []

    for values in worksheet.iter_rows(values_only=True):
        padded = list(values) + [None] * 7
        sector_cell, subsector_cell, segment_cell, name_cell, code_cell = padded[1:6]
        sector_text = str(sector_cell).strip() if sector_cell is not None else ""
        subsector_text = str(subsector_cell).strip() if subsector_cell is not None else ""
        segment_text = str(segment_cell).strip() if segment_cell is not None else ""
        name = str(name_cell).strip() if name_cell is not None else ""
        code = str(code_cell).strip().upper() if code_cell is not None else ""

        if sector_text in {"SECTOR", "ECONOMIC SECTOR"}:
            continue
        if sector_text:
            current_sector = sector_text
        if subsector_text:
            current_subsector = subsector_text
        if segment_text:
            current_segment = segment_text
        if not name or not code or code == "CODE":
            continue

        sector = normalize_sector(B3_ENGLISH_SECTOR_MAP.get(current_sector, ""), "Stock")
        if not sector:
            continue
        rows.append(
            B3SectorRow(
                code=code,
                name=name,
                sector_pt=current_sector,
                sector=sector,
                subsector=current_subsector,
                segment=current_segment,
            )
        )
    return rows


def ticker_root(ticker: str) -> str:
    match = re.match(r"^[A-Z]+", ticker.strip().upper())
    return match.group(0)[:4] if match else ticker.strip().upper()[:4]


def load_target_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    return [
        row
        for row in rows
        if row.get("exchange") == "B3"
        and row.get("asset_type") == "Stock"
        and not row.get("stock_sector", "").strip()
    ]


def evaluate_rows(targets: list[dict[str, str]], b3_rows: list[B3SectorRow]) -> list[dict[str, Any]]:
    by_code: dict[str, list[B3SectorRow]] = {}
    for row in b3_rows:
        by_code.setdefault(row.code, []).append(row)

    results: list[dict[str, Any]] = []
    for target in targets:
        candidates = by_code.get(ticker_root(target["ticker"]), [])
        base = {
            "ticker": target["ticker"],
            "exchange": target["exchange"],
            "asset_type": target["asset_type"],
            "name": target["name"],
            "b3_code": ticker_root(target["ticker"]),
            "b3_name": "",
            "b3_sector_pt": "",
            "b3_subsector": "",
            "b3_segment": "",
            "sector_update": "",
        }
        if not candidates:
            results.append({**base, "decision": "no_b3_code_match"})
            continue
        sectors = {candidate.sector for candidate in candidates}
        if len(sectors) != 1:
            results.append({**base, "decision": "ambiguous_b3_code_match"})
            continue
        candidate = candidates[0]
        results.append(
            {
                **base,
                "b3_name": candidate.name,
                "b3_sector_pt": candidate.sector_pt,
                "b3_subsector": candidate.subsector,
                "b3_segment": candidate.segment,
                "sector_update": candidate.sector,
                "decision": "accept",
            }
        )
    return results


def build_metadata_updates(results: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [
        {
            "ticker": result["ticker"],
            "exchange": result["exchange"],
            "field": "stock_sector",
            "decision": "update",
            "proposed_value": result["sector_update"],
            "confidence": "0.86",
            "reason": (
                "Official B3 sector-classification ZIP mapped the B3 issuer code root to a canonical stock_sector; "
                "accepted only when the B3 code root matched uniquely and resolved to one canonical sector. "
                f"Source: {B3_SECTOR_ZIP_URL}"
            ),
        }
        for result in results
        if result.get("decision") == "accept"
    ]


def write_report_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=REPORT_FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in REPORT_FIELDNAMES})


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Backfill B3 stock sectors from the official B3 sector classification ZIP.")
    parser.add_argument("--tickers-csv", type=Path, default=TICKERS_CSV)
    parser.add_argument("--json-out", type=Path, default=DEFAULT_REPORT_JSON)
    parser.add_argument("--csv-out", type=Path, default=DEFAULT_REPORT_CSV)
    parser.add_argument("--metadata-updates-csv", type=Path, default=DEFAULT_METADATA_UPDATES_CSV)
    parser.add_argument("--source-zip", type=Path)
    parser.add_argument("--source-xlsx", type=Path)
    parser.add_argument("--timeout-seconds", type=float, default=30.0)
    parser.add_argument("--apply", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    zip_bytes = args.source_zip.read_bytes() if args.source_zip else download_b3_sector_zip(B3_SECTOR_ZIP_URL, args.timeout_seconds)
    xlsx_bytes = (
        args.source_xlsx.read_bytes()
        if args.source_xlsx
        else download_b3_industry_xlsx(B3_INDUSTRY_XLSX_URL, args.timeout_seconds)
    )
    b3_rows = [
        *load_b3_sector_rows(extract_workbook(zip_bytes)),
        *load_b3_industry_download_rows(xlsx_bytes),
    ]
    targets = load_target_rows(args.tickers_csv)
    results = evaluate_rows(targets, b3_rows)
    updates = build_metadata_updates(results)
    if args.apply and updates:
        merge_metadata_updates(args.metadata_updates_csv, updates)

    args.json_out.parent.mkdir(parents=True, exist_ok=True)
    args.json_out.write_text(
        json.dumps([row for row in results if row["decision"] == "accept"], indent=2, ensure_ascii=False, sort_keys=True),
        encoding="utf-8",
    )
    write_report_csv(args.csv_out, results)

    print(
        json.dumps(
            {
                "accepted_sector_updates": len(updates),
                "applied": args.apply,
                "b3_classification_rows": len(b3_rows),
                "candidates": len(targets),
                "csv_out": display_path(args.csv_out),
                "decision_counts": dict(Counter(row["decision"] for row in results)),
                "json_out": display_path(args.json_out),
            },
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
