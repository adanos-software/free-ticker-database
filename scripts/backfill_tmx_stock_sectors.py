from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import urllib.request
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.backfill_yahoo_generic_etf_names import merge_metadata_updates
from scripts.rebuild_dataset import TICKERS_CSV, normalize_sector


TMX_ISSUERS_XLSX_URL = "https://www.tsx.com/en/resource/571"
DEFAULT_OUTPUT_DIR = ROOT / "data" / "tmx_verification"
DEFAULT_REPORT_JSON = DEFAULT_OUTPUT_DIR / "stock_sector_backfill.json"
DEFAULT_REPORT_CSV = DEFAULT_OUTPUT_DIR / "stock_sector_backfill.csv"
DEFAULT_METADATA_UPDATES_CSV = ROOT / "data" / "review_overrides" / "metadata_updates.csv"

TMX_STOCK_SECTOR_MAP = {
    "Comm. & Media": "Communication Services",
    "Financial Services": "Financials",
    "Industrial Products & Services": "Industrials",
    "Life Sciences": "Health Care",
    "Mining": "Materials",
    "Oil & Gas": "Energy",
    "Real Estate": "Real Estate",
    "Technology": "Information Technology",
    "Utilities & Pipelines": "Utilities",
}

REPORT_FIELDNAMES = [
    "ticker",
    "exchange",
    "asset_type",
    "name",
    "tmx_symbol",
    "tmx_name",
    "tmx_exchange",
    "tmx_sector",
    "tmx_subsector",
    "sector_update",
    "decision",
]


@dataclass(frozen=True)
class TmxSectorRow:
    exchange: str
    symbol: str
    name: str
    sector: str
    subsector: str


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def download_tmx_issuers_xlsx(url: str, timeout_seconds: float) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "free-ticker-database/3.0"})
    with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
        return response.read()


def clean_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " ")).strip()


def normalize_tmx_symbol(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.strip().upper())


def ticker_root_symbol(value: str) -> str:
    ticker = value.strip().upper()
    root = re.split(r"[.-]", ticker, maxsplit=1)[0]
    return normalize_tmx_symbol(root)


def normalize_tmx_sector(value: str) -> str:
    return normalize_sector(TMX_STOCK_SECTOR_MAP.get(value.strip(), ""), "Stock")


def load_tmx_sector_rows(workbook_bytes: bytes) -> list[TmxSectorRow]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    rows: list[TmxSectorRow] = []
    for worksheet in workbook.worksheets:
        headers: list[str] | None = None
        for values in worksheet.iter_rows(values_only=True):
            cleaned = [clean_header(value) for value in values]
            if "Exchange" in cleaned and "Name" in cleaned and "Root Ticker" in cleaned and "Sector" in cleaned:
                headers = cleaned
                continue
            if headers is None:
                continue
            record = {header: cleaned[index] if index < len(cleaned) else "" for index, header in enumerate(headers)}
            exchange = record.get("Exchange", "").strip().upper()
            symbol = record.get("Root Ticker", "").strip().upper()
            name = record.get("Name", "").strip()
            sector = record.get("Sector", "").strip()
            subsector = record.get("Sub Sector", "") or record.get("Sub-Sector", "")
            if exchange not in {"TSX", "TSXV"} or not symbol or not sector:
                continue
            rows.append(TmxSectorRow(exchange=exchange, symbol=symbol, name=name, sector=sector, subsector=subsector))
    return rows


def load_target_rows(path: Path, exchanges: set[str]) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    return [
        row
        for row in rows
        if row.get("exchange") in exchanges
        and row.get("asset_type") == "Stock"
        and not row.get("stock_sector", "").strip()
    ]


def index_tmx_rows(rows: list[TmxSectorRow]) -> dict[tuple[str, str], list[TmxSectorRow]]:
    indexed: dict[tuple[str, str], list[TmxSectorRow]] = {}
    for row in rows:
        indexed.setdefault((row.exchange, normalize_tmx_symbol(row.symbol)), []).append(row)
        indexed.setdefault((row.exchange, ticker_root_symbol(row.symbol)), []).append(row)
    return indexed


def evaluate_row(row: dict[str, str], matches: list[TmxSectorRow]) -> dict[str, Any]:
    base = {
        "ticker": row.get("ticker", ""),
        "exchange": row.get("exchange", ""),
        "asset_type": row.get("asset_type", ""),
        "name": row.get("name", ""),
        "tmx_symbol": "",
        "tmx_name": "",
        "tmx_exchange": "",
        "tmx_sector": "",
        "tmx_subsector": "",
        "sector_update": "",
    }
    if not matches:
        return {**base, "decision": "no_tmx_symbol_match"}

    sectors = {normalize_tmx_sector(match.sector) for match in matches}
    sectors.discard("")
    source_sectors = {match.sector for match in matches}
    if len(sectors) != 1:
        match = matches[0]
        return {
            **base,
            "tmx_symbol": match.symbol,
            "tmx_name": match.name,
            "tmx_exchange": match.exchange,
            "tmx_sector": "|".join(sorted(source_sectors)),
            "tmx_subsector": match.subsector,
            "decision": "unsupported_or_ambiguous_tmx_sector",
        }

    match = matches[0]
    return {
        **base,
        "tmx_symbol": match.symbol,
        "tmx_name": match.name,
        "tmx_exchange": match.exchange,
        "tmx_sector": "|".join(sorted(source_sectors)),
        "tmx_subsector": match.subsector,
        "sector_update": next(iter(sectors)),
        "decision": "accept",
    }


def evaluate_rows(target_rows: list[dict[str, str]], source_rows: list[TmxSectorRow]) -> list[dict[str, Any]]:
    indexed = index_tmx_rows(source_rows)
    results: list[dict[str, Any]] = []
    for row in target_rows:
        key = (row.get("exchange", ""), ticker_root_symbol(row.get("ticker", "")))
        results.append(evaluate_row(row, indexed.get(key, [])))
    return results


def build_metadata_updates(results: list[dict[str, Any]]) -> list[dict[str, str]]:
    updates: list[dict[str, str]] = []
    for result in results:
        if result.get("decision") != "accept":
            continue
        updates.append(
            {
                "ticker": result["ticker"],
                "exchange": result["exchange"],
                "field": "stock_sector",
                "decision": "update",
                "proposed_value": result["sector_update"],
                "confidence": "0.82",
                "reason": (
                    "Official TMX TSX/TSXV issuer workbook supplied a sector mapped to canonical stock_sector; accepted "
                    "after exact exchange and root-ticker match with a single supported TMX sector."
                ),
            }
        )
    return updates


def write_report_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=REPORT_FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in REPORT_FIELDNAMES})


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Backfill TSX/TSXV stock sectors from the official TMX issuer workbook.")
    parser.add_argument("--source-xlsx", type=Path)
    parser.add_argument("--json-out", type=Path, default=DEFAULT_REPORT_JSON)
    parser.add_argument("--csv-out", type=Path, default=DEFAULT_REPORT_CSV)
    parser.add_argument("--metadata-updates-csv", type=Path, default=DEFAULT_METADATA_UPDATES_CSV)
    parser.add_argument("--exchange", action="append", choices=["TSX", "TSXV"])
    parser.add_argument("--timeout-seconds", type=float, default=30.0)
    parser.add_argument("--apply", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    exchanges = set(args.exchange or ["TSX", "TSXV"])
    workbook_bytes = (
        args.source_xlsx.read_bytes()
        if args.source_xlsx
        else download_tmx_issuers_xlsx(TMX_ISSUERS_XLSX_URL, args.timeout_seconds)
    )
    source_rows = load_tmx_sector_rows(workbook_bytes)
    target_rows = load_target_rows(TICKERS_CSV, exchanges)
    results = evaluate_rows(target_rows, source_rows)
    updates = build_metadata_updates(results)

    args.json_out.parent.mkdir(parents=True, exist_ok=True)
    args.json_out.write_text(json.dumps(results, indent=2), encoding="utf-8")
    write_report_csv(args.csv_out, results)
    if args.apply and updates:
        merge_metadata_updates(args.metadata_updates_csv, updates)

    print(
        json.dumps(
            {
                "accepted_sector_updates": len(updates),
                "applied": args.apply,
                "candidates": len(target_rows),
                "csv_out": display_path(args.csv_out),
                "decision_counts": dict(Counter(result["decision"] for result in results)),
                "exchanges": sorted(exchanges),
                "json_out": display_path(args.json_out),
                "source_rows": len(source_rows),
            },
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
