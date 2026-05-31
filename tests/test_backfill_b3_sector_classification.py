import csv
import json
from io import BytesIO
from zipfile import ZipFile

import scripts.backfill_b3_sector_classification as b3_sector
from openpyxl import Workbook

from scripts.backfill_b3_sector_classification import (
    B3_SECTOR_ZIP_URL,
    apply_gate_context_for,
    evaluate_rows,
    extract_workbook,
    load_b3_industry_download_rows,
    load_b3_sector_rows,
    main,
    official_source_context_for,
    sector_review_context_for,
    ticker_root,
    verification_evidence_required_for,
)


def make_b3_zip() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SETOR ECONÔMICO", "SUBSETOR", "SEGMENTO", "LISTAGEM"])
    worksheet.append(["Utilidade Pública", "Energia Elétrica", "Energia Elétrica", None])
    worksheet.append([None, None, "COPEL", "CPLE"])
    worksheet.append(["Outros", "Outros", "Outros", None])
    worksheet.append([None, None, "IGNORED", "IGNR"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)

    zip_bytes = BytesIO()
    with ZipFile(zip_bytes, "w") as archive:
        archive.writestr("ClassifSetorial.xlsx", workbook_bytes.getvalue())
    return zip_bytes.getvalue()


def make_b3_download_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append([None, "SECTOR", "SUBSECTOR", "SEGMENT", "ISSUER", None, None])
    worksheet.append([None, None, None, None, "TRADING NAME", "CODE", "TRADING SEGMENT"])
    worksheet.append([None, "Utilities", "Electric Power", "Electric Power", "AXIA ENERGIA", "AXIA", "Novo Mercado"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    return workbook_bytes.getvalue()


def test_load_b3_sector_rows_maps_official_sector_and_skips_outros():
    rows = load_b3_sector_rows(extract_workbook(make_b3_zip()))

    assert len(rows) == 1
    assert rows[0].code == "CPLE"
    assert rows[0].sector == "Utilities"


def test_load_b3_industry_download_rows_maps_current_official_file():
    rows = load_b3_industry_download_rows(make_b3_download_workbook())

    assert len(rows) == 1
    assert rows[0].code == "AXIA"
    assert rows[0].sector == "Utilities"


def test_evaluate_rows_accepts_unique_b3_code_root():
    result = evaluate_rows(
        [
            {
                "ticker": "CPLE99",
                "exchange": "B3",
                "asset_type": "Stock",
                "name": "CIA PARANAENSE DE ENERGIA - COPEL",
            }
        ],
        load_b3_sector_rows(extract_workbook(make_b3_zip())),
    )[0]

    assert result["decision"] == "accept"
    assert result["listing_key"] == "B3::CPLE99"
    assert result["sector_update"] == "Utilities"
    assert result["source_url"] == B3_SECTOR_ZIP_URL
    assert result["verification_evidence_required"] == (
        "official_b3_sector_classification_row_with_unique_code_root_and_single_canonical_sector"
    )
    assert result["official_source_context"] == official_source_context_for(result)
    assert result["sector_review_context"] == sector_review_context_for(result)
    assert result["apply_gate_context"] == apply_gate_context_for(result)


def test_ticker_root_removes_b3_suffix_without_losing_alphanumeric_issuer_code():
    assert ticker_root("CPLE3") == "CPLE"
    assert ticker_root("AALR13") == "AALR"
    assert ticker_root("A6OP3") == "A6OP"
    assert ticker_root("2WAV3") == "2WAV"


def test_evaluate_rows_handles_alphanumeric_b3_code_root():
    result = evaluate_rows(
        [
            {
                "ticker": "A6OP3",
                "exchange": "B3",
                "asset_type": "Stock",
                "name": "ACESSOPAR INVESTIMENTOS E PARTICIPAÇÕES S.A.",
            }
        ],
        [
            *load_b3_sector_rows(extract_workbook(make_b3_zip())),
            *load_b3_industry_download_rows(make_b3_download_workbook()),
        ],
    )[0]

    assert result["b3_code"] == "A6OP"
    assert result["decision"] == "no_b3_code_match"
    assert result["verification_evidence_required"] == (
        "official_b3_sector_classification_code_root_match_required_before_sector_update"
    )
    assert result["official_source_context"] == official_source_context_for(result)
    assert result["sector_review_context"] == sector_review_context_for(result)
    assert result["apply_gate_context"] == apply_gate_context_for(result)


def test_evaluate_rows_rejects_missing_code_root():
    result = evaluate_rows(
        [
            {
                "ticker": "ZZZZ3",
                "exchange": "B3",
                "asset_type": "Stock",
                "name": "Unknown",
            }
        ],
        load_b3_sector_rows(extract_workbook(make_b3_zip())),
    )[0]

    assert result["decision"] == "no_b3_code_match"


def test_evaluate_rows_accepts_duplicate_sources_with_same_sector():
    rows = [
        *load_b3_sector_rows(extract_workbook(make_b3_zip())),
        *load_b3_industry_download_rows(make_b3_download_workbook()),
    ]
    result = evaluate_rows(
        [
            {
                "ticker": "CPLE99",
                "exchange": "B3",
                "asset_type": "Stock",
                "name": "CIA PARANAENSE DE ENERGIA - COPEL",
            }
        ],
        [*rows, rows[0]],
    )[0]

    assert result["decision"] == "accept"
    assert result["sector_update"] == "Utilities"


def test_verification_evidence_required_for_documents_ambiguous_b3_code_matches():
    assert verification_evidence_required_for("ambiguous_b3_code_match") == (
        "manual_b3_code_root_review_required_before_sector_update"
    )


def test_main_records_unavailable_official_source_and_uses_remaining_b3_source(tmp_path, monkeypatch, capsys):
    tickers_csv = tmp_path / "tickers.csv"
    with tickers_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["ticker", "exchange", "asset_type", "name", "stock_sector"])
        writer.writeheader()
        writer.writerow(
            {
                "ticker": "AXIA3",
                "exchange": "B3",
                "asset_type": "Stock",
                "name": "AXIA ENERGIA",
                "stock_sector": "",
            }
        )

    def fail_zip(url, timeout_seconds):
        raise RuntimeError("temporary official source outage")

    monkeypatch.setattr(b3_sector, "download_b3_sector_zip", fail_zip)
    monkeypatch.setattr(b3_sector, "download_b3_industry_xlsx", lambda url, timeout_seconds: make_b3_download_workbook())
    json_out = tmp_path / "report.json"
    csv_out = tmp_path / "report.csv"

    main(["--tickers-csv", str(tickers_csv), "--json-out", str(json_out), "--csv-out", str(csv_out)])

    printed = json.loads(capsys.readouterr().out)
    payload = json.loads(json_out.read_text(encoding="utf-8"))
    rows = list(csv.DictReader(csv_out.open(newline="", encoding="utf-8")))
    assert printed["source_fetch_errors"]["b3_sector_zip"].startswith("RuntimeError:")
    assert payload["summary"]["source_fetch_errors"] == printed["source_fetch_errors"]
    assert payload["summary"]["accepted_sector_updates"] == 1
    assert payload["rows"][0]["official_source_context"] == official_source_context_for(payload["rows"][0])
    assert rows[0]["verification_evidence_required"] == (
        "official_b3_sector_classification_row_with_unique_code_root_and_single_canonical_sector"
    )
    assert rows[0]["apply_gate_context"] == apply_gate_context_for(rows[0])
